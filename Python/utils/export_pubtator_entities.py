"""Export distinct Gene entities from the AXIOM corpus PubTator index to Excel.

Reads the pubtator_entities table from axiom.db, aggregates Gene-typed rows by
normalized NCBI Gene ID, and writes one row per distinct gene to a timestamped
.xlsx file. Each output row carries the canonical gene symbol, NCBI Gene ID,
the comma-separated set of surface forms found in the corpus, and the count of
distinct source documents in which the gene appears.

Usage:
    python export_pubtator_entities.py
"""

from __future__ import annotations

import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATABASE_PATH = PROJECT_ROOT / "lib" / "data" / "axiom.db"
EXPORT_DIR = PROJECT_ROOT / "export"

HEADERS = ["gene_symbol", "ncbi_gene_id", "surface_forms", "paper_count"]

QUERY = """
SELECT
    MAX(normalized_name)            AS gene_symbol,
    normalized_id                   AS ncbi_gene_id,
    GROUP_CONCAT(DISTINCT mention)  AS surface_forms,
    COUNT(DISTINCT source_id)       AS paper_count
FROM pubtator_entities
WHERE entity_type = 'Gene'
GROUP BY normalized_id
ORDER BY paper_count DESC, gene_symbol ASC
"""


def fetch_genes(database_path: Path) -> list[tuple]:
    """Run the aggregation query against the corpus index database."""
    if not database_path.exists():
        raise FileNotFoundError(f"Database not found: {database_path}")
    with sqlite3.connect(database_path) as conn:
        return conn.execute(QUERY).fetchall()


def normalize_surface_forms(raw: str | None) -> str:
    """Deduplicate and sort the comma-separated surface form list from SQLite."""
    if not raw:
        return ""
    parts = sorted({segment.strip() for segment in raw.split(",") if segment.strip()})
    return ", ".join(parts)


def write_workbook(rows: list[tuple], output_path: Path) -> None:
    """Write the aggregated gene rows to an Excel workbook."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PubTator Genes"

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    left_align = Alignment(horizontal="left", vertical="top")
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = left_align

    for gene_symbol, ncbi_gene_id, surface_forms, paper_count in rows:
        sheet.append([
            gene_symbol or "",
            ncbi_gene_id,
            normalize_surface_forms(surface_forms),
            paper_count,
        ])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if cell.column_letter == "C":
                cell.alignment = wrap_align
            else:
                cell.alignment = left_align

    sheet.freeze_panes = "A2"

    column_widths = {"A": 16, "B": 14, "C": 60, "D": 14}
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    last_column_letter = get_column_letter(len(HEADERS))
    sheet.auto_filter.ref = f"A1:{last_column_letter}{len(rows) + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    rows = fetch_genes(DATABASE_PATH)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = EXPORT_DIR / f"pubtator_genes_{timestamp}.xlsx"
    write_workbook(rows, output_path)
    print(f"Wrote {len(rows)} distinct genes to {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
